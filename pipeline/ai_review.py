#!/usr/bin/env python3
# Last updated: 2026-07-22 · AI Review layer for MeatCODE data snapshots.
"""
MeatCODE — honest AI review of a raw data snapshot.

Adds ONE extra sheet, "AI Review", to a snapshot .xlsx produced by export_snapshot.py.
The raw sheets stay untouched (still raw, no interpretation). This sheet is the OPINION
layer, clearly labelled as such: an AI model reads the filtration + tagging of every table
and gives a straightforward, fully honest opinion on
    (1) validity      — is the data real, consistent, trustworthy?
    (2) what's missing — empty fields, absent enrichment, coverage gaps
    (3) better tagging — concrete controlled-vocabulary / schema / tagging fixes

How it works:
  * A deterministic Python profiler computes the factual backbone for each sheet
    (row count, per-column fill rate, distinct counts, fully-empty columns, constant
    columns, partial tag coverage, top values). These numbers are never guessed.
  * That profile + a few trimmed sample rows are handed to Claude, which writes the
    narrative opinion grounded in those numbers.
  * If the Anthropic API is unreachable or the key is missing, the sheet is still
    written with the deterministic coverage facts plus a clear note — it never fails
    the snapshot and never leaves an empty sheet.

Two ways to supply the opinion:
  * AUTO (self-contained): the script calls Claude via the Anthropic API. Good for running
    the tool standalone/manually. Note a full review can take ~60s to generate, so give it
    a generous timeout.
  * INJECTED (--review-json): an AI model that has already read the data (e.g. the agent
    running the scheduled audit) supplies the opinion as JSON; the script just embeds it.
    This is the primary path for the recurring audit because it is fast and never depends
    on a slow in-process API round-trip. Either way, "an AI model went through the data".

Run:
    python3 pipeline/ai_review.py path/to/snapshot.xlsx           # review one file (AUTO/API)
    python3 pipeline/ai_review.py --all                           # review every data/snapshots/*.xlsx (AUTO)
    python3 pipeline/ai_review.py --all --force                   # re-review even if a sheet exists
    python3 pipeline/ai_review.py --all --review-json rev.json \
            --model-label "claude-fable-5 (agent-authored)"       # INJECT a pre-written opinion

Reads ANTHROPIC_API_KEY from meatCODE/.env (or environment). Model via AI_REVIEW_MODEL
(default: claude-sonnet-5). Needs: pip install openpyxl
"""
from __future__ import annotations

import argparse
import datetime as _dt
import glob
import hashlib
import json
import os
import urllib.error
import urllib.request
from collections import Counter
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
SHEET_NAME = "AI Review"
RAW_SHEETS = ["Sources", "Experts", "Molecules", "Odours", "Organizations"]
DEFAULT_MODEL = os.environ.get("AI_REVIEW_MODEL", "claude-sonnet-5")
API_URL = "https://api.anthropic.com/v1/messages"

# Short domain context so the model judges the data against what MeatCODE actually needs.
DOMAIN_CONTEXT = (
    "MeatCODE is a flavor & aroma research database for cultivated and plant-based meat R&D. "
    "Sources = scientific literature (mostly from Europe PMC / Dimensions ingestion). "
    "Experts = researchers to potentially map and reach out to. "
    "Molecules = aroma/flavor compounds. Odours = sensory descriptors. "
    "Organizations = companies / NGOs / institutes in the ecosystem. "
    "The point of the tag columns (pathway, method, sensory_descriptor, matrix, compound_class, "
    "study_type, evidence_strength, trust_tier, relevance_llm, etc.) is to make the literature "
    "filterable and to feed a grounded Oracle/RAG assistant."
)


def _env_value(key: str) -> str | None:
    envp = REPO_ROOT / ".env"
    if envp.exists():
        for ln in envp.read_text(encoding="utf-8").splitlines():
            ln = ln.strip()
            if ln.startswith(f"{key}="):
                return ln.split("=", 1)[1].strip().strip('"').strip("'")
    return os.environ.get(key)


# ----------------------------------------------------------------------------- profiling

def _sheet_records(ws):
    """Return (headers, [row-dicts]) for a raw snapshot sheet, skipping any banner."""
    rows = list(ws.iter_rows(values_only=True))
    hidx = next((i for i, r in enumerate(rows) if r and r[0] == "id"), 0)
    headers = [h for h in rows[hidx] if h not in (None, "")]
    recs = []
    for r in rows[hidx + 1:]:
        if not r or r[0] in (None, ""):
            continue
        recs.append(dict(zip(rows[hidx], r)))
    return headers, recs


def profile_sheet(headers, recs):
    n = len(recs)
    cols = {}
    for c in headers:
        vals = [rec.get(c) for rec in recs]
        filled = [v for v in vals if v not in (None, "")]
        distinct = len({str(v) for v in filled})
        info = {"fill": len(filled), "n": n, "distinct": distinct}
        if 0 < distinct <= 12:
            info["top"] = Counter(str(v)[:60] for v in filled).most_common(12)
        cols[c] = info
    empty = [c for c in headers if cols[c]["fill"] == 0]
    constant = [
        (c, cols[c].get("top", [("", 0)])[0][0])
        for c in headers
        if cols[c]["distinct"] == 1 and cols[c]["fill"] == n and n > 0
    ]
    partial = [
        (c, cols[c]["fill"], n)
        for c in headers
        if 0 < cols[c]["fill"] < n
    ]
    return {"n": n, "cols": cols, "empty": empty, "constant": constant, "partial": partial}


def _profile_text(profiles):
    """Compact, model-readable rendering of the deterministic profile."""
    out = []
    for name in RAW_SHEETS:
        p = profiles.get(name)
        if not p:
            continue
        out.append(f"### {name} (n={p['n']})")
        if p["empty"]:
            out.append("  FULLY EMPTY columns (0% filled): " + ", ".join(p["empty"]))
        if p["constant"]:
            out.append("  CONSTANT columns (one value for every row): "
                       + ", ".join(f"{c}={v!r}" for c, v in p["constant"]))
        if p["partial"]:
            out.append("  PARTIAL columns (some rows tagged, some not): "
                       + ", ".join(f"{c} {f}/{n} ({round(100*f/n)}%)" for c, f, n in p["partial"]))
        # low-cardinality value distributions (useful for judging tag vocabularies)
        for c, info in p["cols"].items():
            if "top" in info and info["fill"]:
                dist = ", ".join(f"{v}×{k}" for v, k in info["top"])
                out.append(f"  {c} values: {dist}")
        out.append("")
    return "\n".join(out)


def _samples(ws_records, k=4, maxlen=160, drop=("search_vec",)):
    out = []
    for rec in ws_records[:k]:
        row = {}
        for key, v in rec.items():
            if key in drop or v in (None, ""):
                continue
            s = str(v)
            row[key] = s if len(s) <= maxlen else s[:maxlen] + "…"
        out.append(row)
    return out


# ----------------------------------------------------------------------------- the review

REVIEW_SCHEMA = {
    "headline_verdict": "2-4 blunt sentences: overall state of this snapshot's data quality, filtration and tagging.",
    "tables": {
        name: {"validity": "...", "missing": "...", "tagging": "..."} for name in RAW_SHEETS
    },
    "cross_cutting": ["issue that spans multiple tables or the whole pipeline"],
    "top_fixes": ["ranked, concrete, actionable fix (most impactful first)"],
}


def build_prompt(profile_text, samples):
    return (
        f"{DOMAIN_CONTEXT}\n\n"
        "You are performing an internal data-quality audit of ONE snapshot of this database. "
        "Be a straightforward, fully honest, critical reviewer. Do NOT flatter and do not pad with "
        "praise. Every numeric claim you make must come from the deterministic profile below — do not "
        "invent fill rates. Judge the data the way a skeptical data engineer would.\n\n"
        "For EACH of the five tables, and then across the whole snapshot, assess:\n"
        "  (1) validity  — is what IS there real, internally consistent and trustworthy? Call out "
        "anything that smells like a broken/half-run pipeline (e.g. a scoring or review column that is "
        "the same value for every row, enrichment columns that are entirely empty, suspicious "
        "uniformity, obviously stale timestamps).\n"
        "  (2) missing   — which fields are empty or partial, and what does that make impossible "
        "(e.g. no expert email/country ⇒ cannot do outreach or geographic mapping; no SMILES/CAS ⇒ "
        "molecules are not chemically usable).\n"
        "  (3) tagging   — concretely, how could filtration/tagging be better: controlled vocabularies, "
        "fields that should be normalized/split, weak categories (e.g. a category that is mostly "
        "'Other'), tags that should be backfilled, schema additions.\n\n"
        "Ground everything in these facts:\n\n"
        f"=== DETERMINISTIC PROFILE ===\n{profile_text}\n"
        f"=== SAMPLE ROWS (trimmed) ===\n{json.dumps(samples, ensure_ascii=False, indent=0)[:9000]}\n\n"
        "Return ONLY a single JSON object, no prose around it, matching exactly this shape:\n"
        f"{json.dumps(REVIEW_SCHEMA, ensure_ascii=False, indent=2)}\n"
        "Each 'validity'/'missing'/'tagging' value is a short paragraph (2-5 sentences). "
        "'cross_cutting' and 'top_fixes' are arrays of concise strings. Keep it specific and honest."
    )


def call_claude(prompt, model, api_key, timeout=120):
    body = json.dumps({
        "model": model,
        "max_tokens": 4096,
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")
    req = urllib.request.Request(
        API_URL, data=body,
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
    )
    with urllib.request.urlopen(req, timeout=timeout) as r:
        data = json.load(r)
    text = "".join(part.get("text", "") for part in data.get("content", []))
    # tolerate a stray ```json fence
    t = text.strip()
    if t.startswith("```"):
        t = t.split("\n", 1)[1].rsplit("```", 1)[0]
    start, end = t.find("{"), t.rfind("}")
    return json.loads(t[start:end + 1])


# ----------------------------------------------------------------------------- rendering

def _glance_rows(profiles):
    rows = []
    for name in RAW_SHEETS:
        p = profiles.get(name)
        if not p:
            continue
        ncols = len(p["cols"])
        empty = f"{len(p['empty'])} fully-empty" if p["empty"] else "0 fully-empty"
        parts = [f"{p['n']} rows · {ncols} cols · {empty}"]
        if p["constant"]:
            parts.append("constant: " + ", ".join(c for c, _ in p["constant"]))
        if p["partial"]:
            parts.append("partial tags: " + ", ".join(
                f"{c} {round(100*f/n)}%" for c, f, n in p["partial"]))
        rows.append((name, "  |  ".join(parts)))
    return rows


def write_review_sheet(xlsx_path, review, profiles, model, ok, note=""):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = load_workbook(xlsx_path)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(title=SHEET_NAME)  # appended last

    WINE = "7A1F2B"
    LIGHT = "F3E7E9"
    label_font = Font(bold=True, color="7A1F2B")
    body_align = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 120

    r = 1

    def section(title):
        nonlocal r
        c = ws.cell(row=r, column=1, value=title)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=WINE)
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=WINE)
        r += 1

    def kv(label, value, height_hint=True):
        nonlocal r
        a = ws.cell(row=r, column=1, value=label)
        a.font = label_font
        a.alignment = body_align
        b = ws.cell(row=r, column=2, value=value)
        b.alignment = body_align
        if height_hint and value:
            lines = max(1, (len(str(value)) // 95) + str(value).count("\n") + 1)
            ws.row_dimensions[r].height = min(240, 15 * lines)
        r += 1

    def spacer():
        nonlocal r
        r += 1

    # ---- banner
    title = ws.cell(row=r, column=1, value="AI REVIEW — honest opinion on data filtration & tagging")
    title.font = Font(bold=True, size=13, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=WINE)
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=WINE)
    r += 1
    reviewed = _dt.datetime.now(_dt.timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    kv("Snapshot file:", Path(xlsx_path).name, height_hint=False)
    kv("Reviewed:", reviewed, height_hint=False)
    kv("Reviewer:", f"AI model — {model}" if ok else f"deterministic profile only ({model} unavailable)",
       height_hint=False)
    kv("Note:", "This is an OPINION layer, generated by an AI model. The five raw sheets are unchanged. "
                "Numbers cited come from a deterministic profile of this file.", height_hint=True)
    spacer()

    # ---- deterministic facts
    section("COVERAGE AT A GLANCE  (deterministic facts, not opinion)")
    for name, txt in _glance_rows(profiles):
        kv(name, txt)
    spacer()

    # ---- narrative
    if ok and review:
        section("HEADLINE VERDICT")
        kv("Verdict", review.get("headline_verdict", ""))
        spacer()

        tables = review.get("tables", {}) or {}
        for name in RAW_SHEETS:
            t = tables.get(name)
            if not t:
                continue
            section(name.upper())
            kv("Validity", t.get("validity", ""))
            kv("What's missing", t.get("missing", ""))
            kv("Better tagging", t.get("tagging", ""))
            spacer()

        cc = review.get("cross_cutting", []) or []
        if cc:
            section("CROSS-CUTTING ISSUES")
            for item in cc:
                kv("•", item)
            spacer()

        fixes = review.get("top_fixes", []) or []
        if fixes:
            section("TOP FIXES  (ranked, most impactful first)")
            for i, item in enumerate(fixes, 1):
                kv(f"{i}.", item)
    else:
        section("AI NARRATIVE UNAVAILABLE")
        kv("Reason", note or "The AI model could not be reached for this run. The deterministic "
                            "coverage facts above still stand; re-run pipeline/ai_review.py to add the narrative.")

    wb.save(xlsx_path)


# ----------------------------------------------------------------------------- driver

def review_file(xlsx_path, model, api_key, force=False, cache=None, injected=None, model_label=None):
    from openpyxl import load_workbook

    xlsx_path = str(xlsx_path)
    wb = load_workbook(xlsx_path, read_only=True)
    have = SHEET_NAME in wb.sheetnames
    profiles, records, hasher = {}, {}, hashlib.md5()
    for name in RAW_SHEETS:
        if name not in wb.sheetnames:
            continue
        headers, recs = _sheet_records(wb[name])
        profiles[name] = profile_sheet(headers, recs)
        records[name] = recs
        for rec in recs:
            hasher.update(repr(sorted((str(k), str(v)) for k, v in rec.items())).encode())
    wb.close()

    if have and not force:
        return "skipped (already has AI Review; use --force to redo)"

    label = model_label or model

    # INJECTED mode: opinion already written by an AI model that read the data.
    if injected is not None:
        write_review_sheet(xlsx_path, injected, profiles, label, ok=True, note="")
        return "reviewed (injected AI opinion)"

    data_hash = hasher.hexdigest()
    profile_text = _profile_text(profiles)
    samples = {name: _samples(records.get(name, [])) for name in RAW_SHEETS if name in records}

    review, ok, note = None, False, ""
    if cache is not None and data_hash in cache:
        review, ok, note = cache[data_hash]
        note = "(identical data to an already-reviewed snapshot; reusing that review)"
    elif not api_key:
        note = "ANTHROPIC_API_KEY not found."
    else:
        try:
            review = call_claude(build_prompt(profile_text, samples), model, api_key)
            ok = True
        except urllib.error.HTTPError as e:
            note = f"Anthropic API HTTP {e.code}: {e.read().decode()[:200]}"
        except Exception as e:  # noqa: BLE001
            note = f"{type(e).__name__}: {e}"
        if cache is not None and ok:
            cache[data_hash] = (review, ok, "")

    write_review_sheet(xlsx_path, review, profiles, label, ok, note)
    return "reviewed (AI narrative)" if ok else f"written (facts only) — {note}"


def main() -> int:
    ap = argparse.ArgumentParser(description="Append an honest AI Review sheet to snapshot(s).")
    ap.add_argument("path", nargs="?", help="Snapshot .xlsx to review.")
    ap.add_argument("--all", action="store_true", help="Review every data/snapshots/*.xlsx.")
    ap.add_argument("--force", action="store_true", help="Re-review even if an AI Review sheet exists.")
    ap.add_argument("--model", default=DEFAULT_MODEL, help=f"Anthropic model (default {DEFAULT_MODEL}).")
    ap.add_argument("--review-json", default=None,
                    help="Path to a pre-written review JSON to embed (skips the API call).")
    ap.add_argument("--model-label", default=None,
                    help="Reviewer label to show in the sheet when injecting (e.g. the authoring model).")
    ap.add_argument("--profile", action="store_true",
                    help="Print the deterministic profile (fill rates, empty/constant/partial columns, "
                         "value distributions) for the given file and exit — no API call, no writing. "
                         "Use this to ground a hand-authored review before injecting it.")
    args = ap.parse_args()

    if args.profile:
        if not args.path:
            raise SystemExit("--profile needs a snapshot path.")
        from openpyxl import load_workbook
        wb = load_workbook(args.path, read_only=True)
        profiles = {}
        for name in RAW_SHEETS:
            if name in wb.sheetnames:
                headers, recs = _sheet_records(wb[name])
                profiles[name] = profile_sheet(headers, recs)
        wb.close()
        print(_profile_text(profiles))
        return 0

    api_key = _env_value("ANTHROPIC_API_KEY")
    injected = None
    if args.review_json:
        injected = json.loads(Path(args.review_json).read_text(encoding="utf-8"))
        args.force = True  # injecting means we intend to (re)write the sheet

    if args.all:
        files = sorted(glob.glob(str(REPO_ROOT / "data" / "snapshots" / "meatcode_snapshot_*.xlsx")))
    elif args.path:
        files = [args.path]
    else:
        raise SystemExit("Give a snapshot path or --all.")

    cache: dict = {}
    for f in files:
        status = review_file(f, args.model, api_key, force=args.force, cache=cache,
                             injected=injected, model_label=args.model_label)
        print(f"{Path(f).name}: {status}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

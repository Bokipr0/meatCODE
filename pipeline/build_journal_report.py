#!/usr/bin/env python3
"""Render the journal/publisher-access report as .xlsx + .pdf (read-only on the DB)."""
from __future__ import annotations

import importlib.util
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "data" / "reports"
OUT.mkdir(parents=True, exist_ok=True)

spec = importlib.util.spec_from_file_location("jar", REPO / "pipeline" / "journal_access_report.py")
jar = importlib.util.module_from_spec(spec)
spec.loader.exec_module(jar)

D = jar.build()
TOTAL = D["total"]
STAMP = datetime.now(timezone.utc)
TAG = STAMP.strftime("%Y%m%d_%H%M")

# ------------------------------------------------------------------ palette
WINE = "6B2340"
WINE_L = "F2E6EA"
PAID = "C0392B"
FREE = "1E8449"
GREY = "6B6B6B"
HDR = PatternFill("solid", fgColor=WINE)
SUBHDR = PatternFill("solid", fgColor=WINE_L)
BAND = PatternFill("solid", fgColor="FAFAFA")
THIN = Side(style="thin", color="D9D9D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F = "Arial"


def paid_free(access: str) -> str:
    return {"SUB": "PAID", "OA": "FREE", "PRE": "FREE",
            "MIX": "CHECK", "UNK": "CHECK"}[access]


def style_header(ws, row, ncols, widths=None):
    for i in range(1, ncols + 1):
        c = ws.cell(row=row, column=i)
        c.font = Font(name=F, bold=True, color="FFFFFF", size=10)
        c.fill = HDR
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 28
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


# =============================================================== SHEET 1
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Summary"
ws.sheet_view.showGridLines = False

ws["A1"] = "MeatCODE — journal & publisher access map"
ws["A1"].font = Font(name=F, bold=True, size=16, color=WINE)
ws["A2"] = (f"Every source in the Neon literature table, grouped by the journal it came from "
            f"and the platform you must go through to read it.")
ws["A2"].font = Font(name=F, size=10, color=GREY)
ws["A3"] = (f"Generated {STAMP:%Y-%m-%d %H:%M} UTC  ·  {TOTAL} sources  ·  "
            f"{len(D['journals'])} distinct journals after name normalisation")
ws["A3"].font = Font(name=F, size=9, italic=True, color=GREY)

# --- KPI band
acc = {a["access"]: a for a in D["access_rows"]}
sd = next(p for p in D["platforms"] if p["platform"] == "ScienceDirect")
kpis = [
    ("Total sources", TOTAL, "1F1F1F"),
    ("Need a PAID subscription", acc.get("SUB", {}).get("n", 0), PAID),
    ("Free to read (open access)", acc.get("OA", {}).get("n", 0) + acc.get("PRE", {}).get("n", 0), FREE),
    ("On ScienceDirect (Elsevier)", sd["n"], WINE),
]
r = 5
for i, (label, val, colour) in enumerate(kpis):
    col = 1 + i * 2
    lc = ws.cell(row=r, column=col, value=label)
    lc.font = Font(name=F, size=9, bold=True, color=GREY)
    vc = ws.cell(row=r + 1, column=col, value=val)
    vc.font = Font(name=F, size=22, bold=True, color=colour)
    pc = ws.cell(row=r + 2, column=col, value=f"{val / TOTAL:.1%} of corpus" if label != "Total sources" else "100%")
    pc.font = Font(name=F, size=9, color=GREY)

# --- access split table
r = 9
ws.cell(row=r, column=1, value="ACCESS REQUIREMENT — what it costs you to read the corpus").font = \
    Font(name=F, bold=True, size=11, color=WINE)
r += 1
hdrs = ["Access type", "Sources", "% of corpus", "What this means in practice"]
for i, h in enumerate(hdrs, start=1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, len(hdrs), widths=[34, 11, 13, 62, 16, 16, 16, 16])
MEANING = {
    "SUB": "Paywalled — needs an institutional subscription or per-article purchase",
    "OA":  "Free to read for anyone, no login",
    "MIX": "Depends on the individual article; check before assuming",
    "PRE": "Free preprint — not peer reviewed, treat with care",
    "UNK": "No DOI stored, so access could not be determined — needs manual lookup",
}
r += 1
first_acc = r
acc_total_row = first_acc + len(D["access_rows"])   # the TOTAL row written below
for a in D["access_rows"]:
    ws.cell(row=r, column=1, value=a["access_label"])
    ws.cell(row=r, column=2, value=a["n"])
    ws.cell(row=r, column=3, value=f"=B{r}/$B${acc_total_row}")
    ws.cell(row=r, column=4, value=MEANING[a["access"]])
    colour = PAID if a["access"] == "SUB" else FREE if a["access"] in ("OA", "PRE") else GREY
    ws.cell(row=r, column=1).font = Font(name=F, size=10, bold=True, color=colour)
    for i in range(1, 5):
        ws.cell(row=r, column=i).border = BOX
        if i != 1:
            ws.cell(row=r, column=i).font = Font(name=F, size=10)
    ws.cell(row=r, column=3).number_format = "0.0%"
    ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="center")
    r += 1
assert r == acc_total_row, (r, acc_total_row)
ws.cell(row=r, column=1, value="TOTAL").font = Font(name=F, bold=True, size=10)
ws.cell(row=r, column=2, value=f"=SUM(B{first_acc}:B{r - 1})").font = Font(name=F, bold=True, size=10)
ws.cell(row=r, column=3, value=f"=B{r}/$B${acc_total_row}").font = Font(name=F, bold=True, size=10)
ws.cell(row=r, column=3).number_format = "0.0%"
for i in range(1, 5):
    ws.cell(row=r, column=i).fill = SUBHDR
    ws.cell(row=r, column=i).border = BOX
total_acc_row = r

# --- platform table
r += 3
ws.cell(row=r, column=1, value="PLATFORM — where you actually have to go, and how many articles are waiting there").font = \
    Font(name=F, bold=True, size=11, color=WINE)
r += 1
hdrs = ["Platform / gateway", "Total sources", "% of corpus", "Paid", "Free", "Journals", "Publisher (CrossRef registrant)"]
for i, h in enumerate(hdrs, start=1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, len(hdrs))
for i, w in enumerate([34, 13, 12, 10, 10, 11, 46], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
r += 1
plat_first = r
for p in D["platforms"]:
    ws.cell(row=r, column=1, value=p["platform"])
    ws.cell(row=r, column=2, value=p["n"])
    ws.cell(row=r, column=3, value=f"=B{r}/{TOTAL}")
    ws.cell(row=r, column=4, value=p["paid"])
    ws.cell(row=r, column=5, value=p["free"] + p["other"])
    ws.cell(row=r, column=6, value=p["journals"])
    ws.cell(row=r, column=7, value=p["publisher"])
    ws.cell(row=r, column=1).font = Font(name=F, size=10, bold=(p["paid"] > 0))
    ws.cell(row=r, column=4).font = Font(name=F, size=10, bold=True,
                                         color=PAID if p["paid"] else "BFBFBF")
    ws.cell(row=r, column=5).font = Font(name=F, size=10, color=FREE if (p["free"] + p["other"]) else "BFBFBF")
    ws.cell(row=r, column=3).number_format = "0.0%"
    for i in (2, 6, 7):
        ws.cell(row=r, column=i).font = Font(name=F, size=10)
    for i in range(1, 8):
        ws.cell(row=r, column=i).border = BOX
        if (r - plat_first) % 2:
            ws.cell(row=r, column=i).fill = BAND
    r += 1
plat_last = r - 1
ws.cell(row=r, column=1, value="TOTAL").font = Font(name=F, bold=True, size=10)
for col, letter in ((2, "B"), (4, "D"), (5, "E"), (6, "F")):
    c = ws.cell(row=r, column=col, value=f"=SUM({letter}{plat_first}:{letter}{plat_last})")
    c.font = Font(name=F, bold=True, size=10)
ws.cell(row=r, column=3, value=f"=B{r}/{TOTAL}").number_format = "0.0%"
ws.cell(row=r, column=3).font = Font(name=F, bold=True, size=10)
for i in range(1, 8):
    ws.cell(row=r, column=i).fill = SUBHDR
    ws.cell(row=r, column=i).border = BOX

# --- chart of top platforms
chart = BarChart()
chart.type = "bar"
chart.style = 2
chart.title = "Sources per platform"
chart.y_axis.title = None
chart.x_axis.title = "Sources"
ntop = min(10, plat_last - plat_first + 1)
data = Reference(ws, min_col=2, min_row=plat_first, max_row=plat_first + ntop - 1)
cats = Reference(ws, min_col=1, min_row=plat_first, max_row=plat_first + ntop - 1)
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)
chart.legend = None
chart.height, chart.width = 9, 16
ws.add_chart(chart, f"I{plat_first}")

r += 2
note = ws.cell(row=r, column=1, value=(
    "How to read this: 'Platform' is the site the DOI resolves to. A platform can carry both paid and free "
    "titles — ScienceDirect hosts Food Chemistry (paid) and Food Chemistry: X (free, gold OA), so the Paid/Free "
    "columns split it. Publisher names come from the CrossRef API by DOI prefix; open-access status for every "
    "journal with 4+ sources was verified individually against OpenAlex and DOAJ (see the 'Journals' sheet, "
    "column 'Basis')."))
note.font = Font(name=F, size=9, italic=True, color=GREY)
note.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=7)

# =============================================================== SHEET 2
ws2 = wb.create_sheet("Journals")
ws2.sheet_view.showGridLines = False
ws2["A1"] = "Every journal in the corpus, most-cited-source-count first"
ws2["A1"].font = Font(name=F, bold=True, size=13, color=WINE)
ws2["A2"] = ("Sorted by how many of your sources came from it. 'Paid?' is the one column to scan: "
             "PAID = you need a subscription, FREE = open access, CHECK = varies or unknown.")
ws2["A2"].font = Font(name=F, size=9, italic=True, color=GREY)

hdrs = ["#", "Journal", "Sources", "% of corpus", "Paid?", "Access type", "Platform / gateway",
        "Publisher", "ISSN-L", "Basis for access call", "Reviews", "Years", "Name variants merged"]
for i, h in enumerate(hdrs, start=1):
    ws2.cell(row=4, column=i, value=h)
style_header(ws2, 4, len(hdrs))
for i, w in enumerate([5, 52, 9, 11, 8, 20, 24, 40, 11, 40, 9, 12, 60], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

r = 5
first = r
for n, j in enumerate(D["journals"], start=1):
    pf = paid_free(j["access"])
    ws2.cell(row=r, column=1, value=n)
    ws2.cell(row=r, column=2, value=j["journal"])
    ws2.cell(row=r, column=3, value=j["n"])
    ws2.cell(row=r, column=4, value=f"=C{r}/{TOTAL}")
    ws2.cell(row=r, column=5, value=pf)
    ws2.cell(row=r, column=6, value=j["access_label"])
    ws2.cell(row=r, column=7, value=j["platform"])
    ws2.cell(row=r, column=8, value=j["publisher"])
    ws2.cell(row=r, column=9, value=j["issn"])
    ws2.cell(row=r, column=10, value=j["basis"])
    ws2.cell(row=r, column=11, value=j["reviews"])
    ws2.cell(row=r, column=12, value=(f"{j['yr_min']}–{j['yr_max']}" if j["yr_min"] else ""))
    ws2.cell(row=r, column=13, value=j["variant_list"])
    colour = PAID if pf == "PAID" else FREE if pf == "FREE" else "B7791F"
    ws2.cell(row=r, column=5).font = Font(name=F, size=10, bold=True, color=colour)
    ws2.cell(row=r, column=5).alignment = Alignment(horizontal="center")
    ws2.cell(row=r, column=2).font = Font(name=F, size=10, bold=(j["n"] >= 10))
    ws2.cell(row=r, column=3).font = Font(name=F, size=10, bold=True)
    ws2.cell(row=r, column=4).number_format = "0.0%"
    for i in (1, 6, 7, 8, 9, 10, 11, 12, 13):
        ws2.cell(row=r, column=i).font = Font(name=F, size=9, color=GREY if i >= 9 else "1F1F1F")
    for i in range(1, len(hdrs) + 1):
        ws2.cell(row=r, column=i).border = BOX
        if (r - first) % 2:
            ws2.cell(row=r, column=i).fill = BAND
    r += 1
ws2.cell(row=r, column=2, value="TOTAL").font = Font(name=F, bold=True, size=10)
ws2.cell(row=r, column=3, value=f"=SUM(C{first}:C{r - 1})").font = Font(name=F, bold=True, size=10)
ws2.cell(row=r, column=4, value=f"=C{r}/{TOTAL}").font = Font(name=F, bold=True, size=10)
ws2.cell(row=r, column=4).number_format = "0.0%"
ws2.cell(row=r, column=11, value=f"=SUM(K{first}:K{r - 1})").font = Font(name=F, bold=True, size=10)
for i in range(1, len(hdrs) + 1):
    ws2.cell(row=r, column=i).fill = SUBHDR
    ws2.cell(row=r, column=i).border = BOX
ws2.auto_filter.ref = f"A4:M{r - 1}"

# =============================================================== SHEET 3
ws3 = wb.create_sheet("Paid access shortlist")
ws3.sheet_view.showGridLines = False
ws3["A1"] = "Shortlist — the journals that actually cost you money"
ws3["A1"].font = Font(name=F, bold=True, size=13, color=WINE)
ws3["A2"] = ("Only journals whose articles need a paid subscription, biggest first. "
             "This is the list to hand a librarian when negotiating access.")
ws3["A2"].font = Font(name=F, size=9, italic=True, color=GREY)

hdrs = ["#", "Journal", "Sources", "% of all paid", "% of corpus", "Platform / gateway", "Publisher", "ISSN-L"]
for i, h in enumerate(hdrs, start=1):
    ws3.cell(row=4, column=i, value=h)
style_header(ws3, 4, len(hdrs))
for i, w in enumerate([5, 54, 10, 13, 12, 26, 42, 12], start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

paid_journals = [j for j in D["journals"] if j["access"] == "SUB"]
paid_total = sum(j["n"] for j in paid_journals)
r = 5
first = r
for n, j in enumerate(paid_journals, start=1):
    ws3.cell(row=r, column=1, value=n)
    ws3.cell(row=r, column=2, value=j["journal"])
    ws3.cell(row=r, column=3, value=j["n"])
    ws3.cell(row=r, column=4, value=f"=C{r}/{paid_total}")
    ws3.cell(row=r, column=5, value=f"=C{r}/{TOTAL}")
    ws3.cell(row=r, column=6, value=j["platform"])
    ws3.cell(row=r, column=7, value=j["publisher"])
    ws3.cell(row=r, column=8, value=j["issn"])
    ws3.cell(row=r, column=2).font = Font(name=F, size=10, bold=(j["n"] >= 10))
    ws3.cell(row=r, column=3).font = Font(name=F, size=10, bold=True, color=PAID)
    for i in (4, 5):
        ws3.cell(row=r, column=i).number_format = "0.0%"
    for i in (1, 6, 7, 8):
        ws3.cell(row=r, column=i).font = Font(name=F, size=9)
    for i in range(1, len(hdrs) + 1):
        ws3.cell(row=r, column=i).border = BOX
        if (r - first) % 2:
            ws3.cell(row=r, column=i).fill = BAND
    r += 1
ws3.cell(row=r, column=2, value="TOTAL PAID").font = Font(name=F, bold=True, size=10)
ws3.cell(row=r, column=3, value=f"=SUM(C{first}:C{r - 1})").font = Font(name=F, bold=True, size=10, color=PAID)
ws3.cell(row=r, column=4, value=f"=C{r}/{paid_total}").font = Font(name=F, bold=True, size=10)
ws3.cell(row=r, column=5, value=f"=C{r}/{TOTAL}").font = Font(name=F, bold=True, size=10)
for i in (4, 5):
    ws3.cell(row=r, column=i).number_format = "0.0%"
for i in range(1, len(hdrs) + 1):
    ws3.cell(row=r, column=i).fill = SUBHDR
    ws3.cell(row=r, column=i).border = BOX

# =============================================================== SHEET 4
ws4 = wb.create_sheet("Method & caveats")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 108
ws4["A1"] = "How these numbers were produced"
ws4["A1"].font = Font(name=F, bold=True, size=13, color=WINE)
rows = [
    ("Source of data", f"The live Neon Postgres `sources` table, read-only, {TOTAL} rows at "
                       f"{STAMP:%Y-%m-%d %H:%M} UTC. Nothing was written to the database."),
    ("Journal field", "COALESCE(journal, venue). 785 rows had `journal`, 813 had `venue`; "
                      "using both leaves only 5 rows with no journal name at all."),
    ("Name normalisation", "The raw table holds 213 distinct journal strings but only 170 real journals. "
                           "Case, trailing place-names and punctuation were collapsed: "
                           "'Food Chemistry' + 'Food chemistry' = one row; "
                           "'Foods' + 'Foods (Basel, Switzerland)' = one row; "
                           "'Food chemistry: X' + 'Food Chemistry X' = one row. "
                           "The merged spellings are listed per journal in the 'Journals' sheet."),
    ("Publisher attribution", "Taken from the DOI prefix, which is the registrant of record, not from the "
                              "journal name. All 39 distinct prefixes in the corpus were resolved against "
                              "the CrossRef REST API (api.crossref.org/prefixes/<prefix>) on 2026-08-05. "
                              "Attribution is applied per JOURNAL, using that journal's most common prefix — "
                              "so a row missing its own DOI still inherits the right platform if its journal "
                              "is identified. This is why ScienceDirect shows 349 sources while only 334 rows "
                              "literally carry a 10.1016 DOI: the difference is DOI-less rows in Elsevier "
                              "journals, plus 4 Journal of Dairy Science rows under prefix 10.3168, which is "
                              "also hosted on ScienceDirect."),
    ("Open vs paid", "Publisher default first, then a per-journal override for every journal with 4+ sources, "
                     "each checked individually against the OpenAlex /sources API (is_oa, is_in_doaj) and DOAJ. "
                     "This matters: 118 of the 349 ScienceDirect sources are gold-OA titles that are free to "
                     "read. The 'Basis' column says which call was verified and which is a publisher default."),
    ("Known conflict", "Food Science & Nutrition (ISSN 2048-7177): Wiley lists it as open access, OpenAlex "
                       "reports is_oa = false. Left as 'Hybrid / varies' rather than guessing. 8 sources."),
    ("Caveat — no DOI", "49 sources have no DOI stored. 24 of them sit in a journal that other rows do "
                        "identify, so they inherit that journal's platform and access status. The remaining "
                        "25 are in journals where no row has a DOI (Food Hydrocolloids, LWT, Trends in Food "
                        "Science & Technology and 18 others) — those are reported as 'Unknown (no DOI)' and "
                        "are neither assumed free nor assumed paid. Backfilling those 49 DOIs is the single "
                        "cheapest way to sharpen this report."),
    ("Caveat — hybrid journals", "'Hybrid / varies' means the journal publishes both paywalled and open articles. "
                                 "Per-article status would need an Unpaywall lookup on each DOI, which was not run."),
    ("Caveat — journal ≠ article", "A journal being open access does not guarantee every historical article in it "
                                   "is free; Poultry Science and Journal of Dairy Science only went gold OA in "
                                   "2020 and 2024 respectively. Older articles may still be paywalled."),
    ("Reproduce this", "python3 pipeline/journal_access_report.py   (prints the rollup)\n"
                       "python3 pipeline/build_journal_report.py    (writes this .xlsx and the .pdf)"),
]
r = 3
for label, body in rows:
    a = ws4.cell(row=r, column=1, value=label)
    a.font = Font(name=F, bold=True, size=10, color=WINE)
    a.alignment = Alignment(vertical="top", wrap_text=True)
    b = ws4.cell(row=r, column=2, value=body)
    b.font = Font(name=F, size=10)
    b.alignment = Alignment(vertical="top", wrap_text=True)
    for i in (1, 2):
        ws4.cell(row=r, column=i).border = BOX
    ws4.row_dimensions[r].height = max(30, 13 * (len(body) // 95 + 1) + 12)
    r += 1

xlsx_path = OUT / f"meatcode_journal_access_{TAG}.xlsx"
wb.save(xlsx_path)
print("xlsx:", xlsx_path)

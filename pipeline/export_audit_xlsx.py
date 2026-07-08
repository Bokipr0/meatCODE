#!/usr/bin/env python3
# Last updated: 2026-07-08 10:11 UTC · Data Engineer · added --all-sources full-corpus mode (one row per source, taxonomy-relevance-checked) alongside the original single-run export
"""Export a source-audit run OR the whole corpus to a formatted .xlsx workbook.

Two modes:

1. Single-run mode (original, default when --run/no flag given): for a given run in
   `source_audits` (default: the most recent), writes one row per AUDITED source with
   its stored info, the taxonomy tags currently ATTACHED (`source_topics` -> `topics`),
   the judge's verdict + sub-scores, SUGGESTED tags / tag issues, and notes. Two sheets:
     - "Overview": run metadata + verdict counts (COUNTIF formulas) + a column legend.
     - "Audited Sources": per-source table (verdict colour-coded, autofilter, frozen header).
   This is the companion to `audit_sources.py`: the audit writes verdicts to Neon; this
   turns any run into a human-readable / shareable spreadsheet — in particular a
   back-tagging worksheet, since audited legacy rows are typically untagged and the
   "Suggested tags" column is the judge's recommendation for what to tag them with.

2. Full-corpus mode (`--all-sources`): one row per source in ALL of `sources` (not just
   an audited run), combining stored info + attached tags + the taxonomy-bible
   keyword-overlap relevance signal + `relevance_llm` + any prior `source_audits`
   verdict, colour-coded by recommended action. Computed via
   `pipeline/check_relevance_vs_taxonomy.compute_all()` — the SAME function that
   produces `analysis/relevance_check_<date>.md`, so the xlsx and the md always agree.

Usage:
    python3 pipeline/export_audit_xlsx.py [--run RUN_ID] [--out PATH]
    # default: newest run in source_audits -> docs/audits/audit_<date>_<run8>_sources.xlsx

    python3 pipeline/export_audit_xlsx.py --all-sources [--out PATH]
    # default: docs/audits/relevance_check_<date>.xlsx
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))

# Same self-sufficient Neon accessor pattern as audit_sources.py: prefer the repo's
# db.connect, fall back to a direct psycopg2 connection built from .env.
try:
    from db.connect import get_conn            # shared accessor — do NOT reinvent
except Exception:
    def _load_env_once():
        envp = REPO_ROOT / ".env"
        if envp.exists():
            for line in envp.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ.setdefault(k, v.strip().strip('"').strip("'"))

    def get_conn():
        _load_env_once()
        import psycopg2
        url = os.environ.get("DATABASE_URL")
        if not url:
            raise RuntimeError("DATABASE_URL not set (checked .env)")
        return psycopg2.connect(url)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- brand-ish palette (wine / pomegranate) --------------------------------
WINE = "6B1E2E"
WINE_LIGHT = "F3E6E9"
GREY = "6B7280"
VERDICT_FILL = {
    "keep": PatternFill("solid", fgColor="C6EFCE"),
    "review": PatternFill("solid", fgColor="FFEB9C"),
    "quarantine": PatternFill("solid", fgColor="FFC7CE"),
}
VERDICT_FONT = {
    "keep": Font(name="Arial", bold=True, color="1E6B34"),
    "review": Font(name="Arial", bold=True, color="7A5B00"),
    "quarantine": Font(name="Arial", bold=True, color="9C1B2E"),
}
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# --- full-corpus mode: recommended-action palette (--all-sources) ----------
ACTION_LABEL = {
    "keep": "Keep",
    "review": "Review",
    "off_topic_check_first": "Off-topic — check first",
    "off_topic_high_confidence": "Off-topic — high confidence",
}
ACTION_FILL = {
    "keep": PatternFill("solid", fgColor="C6EFCE"),
    "review": PatternFill("solid", fgColor="FFEB9C"),
    "off_topic_check_first": PatternFill("solid", fgColor="FFD9A0"),
    "off_topic_high_confidence": PatternFill("solid", fgColor="FFC7CE"),
}
ACTION_FONT = {
    "keep": Font(name="Arial", bold=True, color="1E6B34"),
    "review": Font(name="Arial", bold=True, color="7A5B00"),
    "off_topic_check_first": Font(name="Arial", bold=True, color="9C5700"),
    "off_topic_high_confidence": Font(name="Arial", bold=True, color="9C1B2E"),
}

# (header, row-key, width, wrap)
COLUMNS = [
    ("#", "_rank", 4, False),
    ("Source ID", "source_id", 9, False),
    ("Verdict", "verdict", 12, False),
    ("Title", "name", 52, True),
    ("Year", "year", 7, False),
    ("Journal / venue", "_journal", 20, True),
    ("Review?", "_review", 8, False),
    ("Citations", "citation_count", 9, False),
    ("DOI", "doi", 22, True),
    ("Authors", "authors", 30, True),
    ("Ingest query", "search_query", 12, False),
    ("Priority score", "priority_score", 10, False),
    ("Relevance (ingest LLM)", "relevance_llm", 11, False),
    ("Audit: tag", "tag_score", 7, False),
    ("Audit: relevance", "relevance_score", 9, False),
    ("Audit: quality", "quality_score", 8, False),
    ("Audit priority", "audit_priority", 9, False),
    ("Attached tags (source_topics)", "_attached", 30, True),
    ("top_keywords", "_topkw", 24, True),
    ("Suggested tags / issues (judge)", "_suggested", 62, True),
    ("Judge notes", "notes", 60, True),
]

SQL = """
  select a.source_id, a.verdict, a.tag_score, a.relevance_score, a.quality_score,
         a.notes, a.tag_issues, a.audit_priority,
         s.name, s.year, s.journal, s.venue, s.doi, s.authors, s.citation_count,
         s.is_review, s.search_query, s.priority_score, s.relevance_llm, s.top_keywords,
         (select string_agg(t.name, '; ' order by t.name)
            from source_topics st join topics t on t.id = st.topic_id
           where st.source_id = a.source_id) as attached_tags
    from source_audits a
    join sources s on s.id = a.source_id
   where a.run_id = %s
   order by a.audit_priority desc nulls last, a.source_id
"""

COLS = ["source_id", "verdict", "tag_score", "relevance_score", "quality_score",
        "notes", "tag_issues", "audit_priority", "name", "year", "journal", "venue",
        "doi", "authors", "citation_count", "is_review", "search_query",
        "priority_score", "relevance_llm", "top_keywords", "attached_tags"]


def latest_run(cur) -> str | None:
    cur.execute("select run_id from source_audits order by audited_at desc limit 1")
    row = cur.fetchone()
    return row[0] if row else None


def fetch_rows(cur, run_id: str) -> list[dict]:
    cur.execute(SQL, (run_id,))
    out = []
    for r in cur.fetchall():
        d = dict(zip(COLS, r))
        d["_journal"] = d.get("journal") or d.get("venue") or ""
        d["_review"] = "Yes" if d.get("is_review") else "No"
        d["_attached"] = d.get("attached_tags") or "— none (untagged) —"
        d["_topkw"] = d.get("top_keywords") or ""
        issues = d.get("tag_issues") or []
        if isinstance(issues, str):
            issues = [issues]
        d["_suggested"] = "\n".join(f"• {x}" for x in issues) if issues else ""
        d["priority_score"] = float(d["priority_score"]) if d.get("priority_score") is not None else None
        d["audit_priority"] = float(d["audit_priority"]) if d.get("audit_priority") is not None else None
        out.append(d)
    for i, d in enumerate(out, 1):
        d["_rank"] = i
    return out


def build_workbook(rows: list[dict], run_id: str) -> Workbook:
    wb = Workbook()

    # ---- Overview sheet ----------------------------------------------------
    ov = wb.active
    ov.title = "Overview"
    ov.sheet_view.showGridLines = False
    counts = {"keep": 0, "review": 0, "quarantine": 0}
    for d in rows:
        counts[d["verdict"]] = counts.get(d["verdict"], 0) + 1
    now = datetime.now(timezone.utc)

    ov["B2"] = "MeatCODE — Source Audit Export"
    ov["B2"].font = Font(name="Arial", bold=True, size=16, color=WINE)
    ov["B3"] = "Information + attached key tags for every source authenticated in this audit run."
    ov["B3"].font = Font(name="Arial", italic=True, color=GREY)

    meta = [
        ("Run ID", run_id),
        ("Generated (UTC)", now.strftime("%Y-%m-%d %H:%M")),
        ("Sources audited", len(rows)),
        ("Dated report", f"docs/audits/ (run {run_id[:8]})"),
    ]
    r = 5
    for label, val in meta:
        ov.cell(r, 2, label).font = Font(name="Arial", bold=True)
        ov.cell(r, 3, val).font = Font(name="Arial")
        r += 1

    r += 1
    ov.cell(r, 2, "Verdicts").font = Font(name="Arial", bold=True, size=12, color=WINE)
    r += 1
    data_last = len(rows) + 1  # header on row 1 of data sheet
    verdict_rows = {
        "keep": ("Keep — on-topic, good quality; no action", "1E6B34"),
        "review": ("Review — borderline / needs a look", "7A5B00"),
        "quarantine": ("Quarantine — off-topic/low quality; confirm before removing", "9C1B2E"),
    }
    for v, (desc, color) in verdict_rows.items():
        ov.cell(r, 2, v.capitalize()).font = Font(name="Arial", bold=True, color=color)
        ov.cell(r, 2).fill = VERDICT_FILL[v]
        # dynamic count via COUNTIF on the data sheet's Verdict column (C)
        ov.cell(r, 3, f'=COUNTIF(\'Audited Sources\'!C2:C{data_last},"{v}")').font = Font(name="Arial", bold=True)
        ov.cell(r, 4, desc).font = Font(name="Arial", color=GREY)
        r += 1

    r += 1
    ov.cell(r, 2, "Column legend").font = Font(name="Arial", bold=True, size=12, color=WINE)
    r += 1
    legend = [
        ("Attached tags (source_topics)", "Canonical taxonomy tags CURRENTLY attached to the source. "
         "\"— none (untagged) —\" means the row is an untagged legacy source (a back-tagging candidate)."),
        ("Suggested tags / issues (judge)", "The audit judge's recommended tags and tagging problems for this "
         "source — use these to back-tag untagged rows."),
        ("Audit: tag / relevance / quality", "The judge's 0–100 sub-scores this run. tag=50 usually means "
         "\"can't assess — no tags stored\"."),
        ("Relevance (ingest LLM)", "The relevance score assigned at ingest time. Compare with Audit: relevance "
         "to see where the audit is stricter than the ingest gate."),
    ]
    for name, desc in legend:
        c = ov.cell(r, 2, name)
        c.font = Font(name="Arial", bold=True)
        c.alignment = Alignment(vertical="top")
        d = ov.cell(r, 3, desc)
        d.font = Font(name="Arial", color="333333")
        d.alignment = Alignment(wrap_text=True, vertical="top")
        ov.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ov.row_dimensions[r].height = 42
        r += 1

    ov.column_dimensions["A"].width = 2
    ov.column_dimensions["B"].width = 30
    ov.column_dimensions["C"].width = 16
    for col in "DEFG":
        ov.column_dimensions[col].width = 22

    # ---- Audited Sources sheet --------------------------------------------
    ws = wb.create_sheet("Audited Sources")
    ws.sheet_view.showGridLines = False
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=WINE)
    for j, (title, _key, width, _wrap) in enumerate(COLUMNS, 1):
        c = ws.cell(1, j, title)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.row_dimensions[1].height = 34

    for i, d in enumerate(rows, start=2):
        for j, (_title, key, _width, wrap) in enumerate(COLUMNS, 1):
            val = d.get(key)
            if key == "year" and val is not None:
                val = str(val)  # years as text, no thousands separator
            c = ws.cell(i, j, val)
            c.font = Font(name="Arial")
            c.alignment = Alignment(wrap_text=wrap, vertical="top",
                                    horizontal="left" if wrap else "center")
            c.border = BORDER
            if key in ("priority_score", "audit_priority"):
                c.number_format = "0.0"
        vcell = ws.cell(i, 3)
        vcell.fill = VERDICT_FILL.get(d["verdict"], PatternFill())
        vcell.font = VERDICT_FONT.get(d["verdict"], Font(name="Arial", bold=True))
        vcell.alignment = Alignment(vertical="center", horizontal="center")
        # row height: scale with the longest wrapped field so notes/suggested show
        longest = max(len(str(d.get("_suggested", ""))), len(str(d.get("notes", ""))),
                      len(str(d.get("name", ""))))
        ws.row_dimensions[i].height = min(220, max(56, (longest // 60 + 1) * 15 + 12))

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
    return wb


# ===========================================================================
# Full-corpus mode (--all-sources): one row per source, taxonomy-relevance-checked
# ===========================================================================
COLUMNS_FULL = [
    ("#", "_rank", 4, False),
    ("Source ID", "id", 9, False),
    ("Recommended action", "_action_label", 24, False),
    ("Title", "name", 50, True),
    ("Year", "year", 7, False),
    ("Journal / venue", "_journal", 20, True),
    ("Review?", "_review", 8, False),
    ("Citations", "citation_count", 9, False),
    ("Abstract?", "_abstract", 9, False),
    ("DOI", "doi", 22, True),
    ("Ingest query", "search_query", 14, False),
    ("Priority score", "priority_score", 10, False),
    ("Relevance (LLM)", "relevance_llm", 10, False),
    ("Taxonomy signal", "taxonomy_signal", 12, False),
    ("Taxonomy hits", "taxonomy_hit_count", 9, False),
    ("Taxonomy branches", "_branches", 24, True),
    ("Taxonomy topics matched", "_hit_names", 38, True),
    ("Attached tags (source_topics)", "_attached", 28, True),
    ("top_keywords", "_topkw", 22, True),
    ("Reconciliation", "reconciliation", 13, False),
    ("Prior audit verdict", "audit_verdict", 12, False),
    ("Prior audit notes", "audit_notes", 48, True),
]


def fetch_full_corpus_rows() -> list[dict]:
    """One row per source in `sources`, via check_relevance_vs_taxonomy.compute_all() —
    the SAME function analysis/relevance_check_<date>.md is built from, so the xlsx and
    the md report never drift apart."""
    from pipeline.check_relevance_vs_taxonomy import compute_all
    conn = get_conn()
    cur = conn.cursor()
    rows = compute_all(cur)
    conn.close()
    for i, d in enumerate(rows, 1):
        d["_rank"] = i
        d["_journal"] = d.get("journal") or d.get("venue") or ""
        d["_review"] = "Yes" if d.get("is_review") else "No"
        d["_abstract"] = "Yes" if d.get("abstract_present") else "No"
        d["_branches"] = "; ".join(d.get("taxonomy_branches") or []) or "—"
        d["_hit_names"] = "; ".join(d.get("taxonomy_hit_names") or []) or "—"
        d["_attached"] = "; ".join(d.get("attached_tags") or []) or "— none (untagged) —"
        d["_topkw"] = d.get("top_keywords") or ""
        d["_action_label"] = ACTION_LABEL.get(d.get("recommended_action"), d.get("recommended_action") or "")
        d["priority_score"] = float(d["priority_score"]) if d.get("priority_score") is not None else None
    return rows


def build_full_corpus_workbook(rows: list[dict]) -> Workbook:
    wb = Workbook()
    now = datetime.now(timezone.utc)

    # ---- Overview sheet -----------------------------------------------------
    ov = wb.active
    ov.title = "Overview"
    ov.sheet_view.showGridLines = False
    n = len(rows)
    citable = sum(1 for d in rows if d.get("abstract_present"))
    tagged = sum(1 for d in rows if d.get("is_tagged"))

    ov["B2"] = "MeatCODE — Full-Corpus Relevance Review"
    ov["B2"].font = Font(name="Arial", bold=True, size=16, color=WINE)
    ov["B3"] = "One row per source: info + attached tags + taxonomy-bible relevance signal + relevance_llm + any prior audit verdict."
    ov["B3"].font = Font(name="Arial", italic=True, color=GREY)

    meta = [
        ("Generated (UTC)", now.strftime("%Y-%m-%d %H:%M")),
        ("Total sources", n),
        ("Citable (abstract + search_vec)", f"{citable} ({100.0*citable/n:.1f}%)" if n else citable),
        ("Tagged (source_topics)", f"{tagged} ({100.0*tagged/n:.1f}%)" if n else tagged),
        ("Untagged", f"{n - tagged} ({100.0*(n-tagged)/n:.1f}%)" if n else n - tagged),
    ]
    r = 5
    for label, val in meta:
        ov.cell(r, 2, label).font = Font(name="Arial", bold=True)
        ov.cell(r, 3, val).font = Font(name="Arial")
        r += 1

    r += 1
    ov.cell(r, 2, "Recommended action (colour-coded on the data sheet)").font = Font(name="Arial", bold=True, size=12, color=WINE)
    r += 1
    data_last = n + 1
    action_rows = {
        "keep": ("Keep — relevance_llm >= 60", "1E6B34"),
        "review": ("Review — relevance_llm 40-59 (tangential)", "7A5B00"),
        "off_topic_check_first": ("Off-topic, check first — LLM<40 but taxonomy matched 2+ topics (disagreement)", "9C5700"),
        "off_topic_high_confidence": ("Off-topic, high confidence — LLM<40, taxonomy agrees (0-1 hits)", "9C1B2E"),
    }
    for key, (desc, color) in action_rows.items():
        label = ACTION_LABEL[key]
        ov.cell(r, 2, label).font = Font(name="Arial", bold=True, color=color)
        ov.cell(r, 2).fill = ACTION_FILL[key]
        ov.cell(r, 3, f'=COUNTIF(\'All Sources\'!C2:C{data_last},"{label}")').font = Font(name="Arial", bold=True)
        ov.cell(r, 4, desc).font = Font(name="Arial", color=GREY)
        r += 1

    r += 1
    ov.cell(r, 2, "Column legend").font = Font(name="Arial", bold=True, size=12, color=WINE)
    r += 1
    legend = [
        ("Taxonomy signal", "off-topic (0 canonical topics matched in name+abstract+top_keywords, unioned with "
         "any already-attached tags) / weak (1 matched) / on-topic (2+ matched). Diagnostic, NOT the verdict — "
         "see Reconciliation."),
        ("Taxonomy hits / branches / topics matched", "The size, branch membership, and names of that matched-topic set."),
        ("Reconciliation", "How the taxonomy signal relates to relevance_llm: agree / borderline (LLM 40-59) / "
         "disagreement (LLM<40 but taxonomy matched 2+ topics — the flagged case) / coverage_gap (LLM>=60 but "
         "zero taxonomy signal and untagged — a back-tagging candidate, not a relevance risk)."),
        ("Attached tags (source_topics)", "Canonical taxonomy tags CURRENTLY attached. \"— none (untagged) —\" "
         "flags a back-tagging candidate."),
        ("Prior audit verdict / notes", "The most recent verdict from the recurring source_audits loop "
         "(pipeline/audit_sources.py), if this source has been audited yet — independent third signal, shown "
         "for cross-check, not blended in."),
    ]
    for name, desc in legend:
        c = ov.cell(r, 2, name)
        c.font = Font(name="Arial", bold=True)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        d = ov.cell(r, 3, desc)
        d.font = Font(name="Arial", color="333333")
        d.alignment = Alignment(wrap_text=True, vertical="top")
        ov.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ov.row_dimensions[r].height = 56
        r += 1

    ov.column_dimensions["A"].width = 2
    ov.column_dimensions["B"].width = 32
    ov.column_dimensions["C"].width = 16
    for col in "DEFG":
        ov.column_dimensions[col].width = 22

    # ---- All Sources sheet --------------------------------------------------
    ws = wb.create_sheet("All Sources")
    ws.sheet_view.showGridLines = False
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=WINE)
    for j, (title, _key, width, _wrap) in enumerate(COLUMNS_FULL, 1):
        c = ws.cell(1, j, title)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.row_dimensions[1].height = 34

    for i, d in enumerate(rows, start=2):
        for j, (_title, key, _width, wrap) in enumerate(COLUMNS_FULL, 1):
            val = d.get(key)
            if key == "year" and val is not None:
                val = str(val)
            c = ws.cell(i, j, val)
            c.font = Font(name="Arial")
            c.alignment = Alignment(wrap_text=wrap, vertical="top",
                                    horizontal="left" if wrap else "center")
            c.border = BORDER
            if key in ("priority_score",):
                c.number_format = "0.0"
        acell = ws.cell(i, 3)
        action_key = d.get("recommended_action")
        acell.fill = ACTION_FILL.get(action_key, PatternFill())
        acell.font = ACTION_FONT.get(action_key, Font(name="Arial", bold=True))
        acell.alignment = Alignment(vertical="center", horizontal="center")
        longest = max(len(str(d.get("_hit_names", ""))), len(str(d.get("audit_notes") or "")),
                      len(str(d.get("name", ""))))
        ws.row_dimensions[i].height = min(160, max(30, (longest // 60 + 1) * 15 + 12))

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS_FULL))}{len(rows) + 1}"
    return wb


def main() -> int:
    ap = argparse.ArgumentParser(description="Export a source_audits run, or the whole corpus, to a formatted xlsx.")
    ap.add_argument("--run", help="run_id to export (default: newest run in source_audits)")
    ap.add_argument("--out", help="output path (default depends on mode — see module docstring)")
    ap.add_argument("--all-sources", action="store_true",
                     help="full-corpus mode: one row per source (info + attached tags + taxonomy "
                          "relevance signal + relevance_llm + any prior audit verdict) instead of one "
                          "audited run")
    args = ap.parse_args()

    now = datetime.now(timezone.utc)

    if args.all_sources:
        rows = fetch_full_corpus_rows()
        if not rows:
            print("!! sources table is empty — nothing to export.")
            return 2
        out = Path(args.out) if args.out else (
            REPO_ROOT / "docs" / "audits" / f"relevance_check_{now:%Y-%m-%d}.xlsx")
        out.parent.mkdir(parents=True, exist_ok=True)
        wb = build_full_corpus_workbook(rows)
        wb.save(out)
        print(f"   wrote {len(rows)} sources -> {out}")
        print(f"   recommended action: " + " · ".join(
            f"{sum(1 for d in rows if d['recommended_action'] == k)} {ACTION_LABEL[k]}"
            for k in ("keep", "review", "off_topic_check_first", "off_topic_high_confidence")))
        return 0

    conn = get_conn()
    cur = conn.cursor()
    run_id = args.run or latest_run(cur)
    if not run_id:
        print("!! no runs found in source_audits — nothing to export.")
        return 2
    rows = fetch_rows(cur, run_id)
    conn.close()
    if not rows:
        print(f"!! run {run_id} has no rows.")
        return 2

    out = Path(args.out) if args.out else (
        REPO_ROOT / "docs" / "audits" / f"audit_{now:%Y-%m-%d}_{run_id[:8]}_sources.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(rows, run_id)
    wb.save(out)
    print(f"   wrote {len(rows)} sources -> {out}")
    print(f"   verdicts: " + " · ".join(f"{sum(1 for d in rows if d['verdict']==v)} {v}"
                                         for v in ("keep", "review", "quarantine")))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

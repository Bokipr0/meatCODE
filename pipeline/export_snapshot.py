#!/usr/bin/env python3
# Last updated: 2026-07-07 11:10 UTC · Data snapshot · raw Neon → 1 xlsx (5 sheets), newest-first, no interpretation
"""
MeatCODE — raw Neon snapshot to Excel.

Pulls the NEWEST rows straight from the live Neon database into ONE .xlsx with five
sheets, exactly as stored — every column, no scoring, no AI commentary, no summaries.

  Sheet 1  Sources        newest 50 rows of  sources
  Sheet 2  Experts        newest 20 rows of  experts
  Sheet 3  Molecules      newest 20 rows of  molecules
  Sheet 4  Odours         newest 20 rows of  odours
  Sheet 5  Organizations  newest 20 rows of  organizations

"Newest" = ORDER BY created_at DESC NULLS LAST, id DESC (all five tables have created_at).
The first sheet (Sources) carries a small provenance banner at the top: the exact
date+time the snapshot was taken (UTC and local) and the origin (Neon host / database).

Run:
    python3 pipeline/export_snapshot.py                 # writes data/snapshots/meatcode_snapshot_<UTCstamp>.xlsx
    python3 pipeline/export_snapshot.py --out FILE.xlsx # custom path

Reads DATABASE_URL from meatCODE/.env. Needs: pip install openpyxl psycopg2-binary
"""
from __future__ import annotations

import argparse
import datetime as _dt
import json
from decimal import Decimal
from pathlib import Path
from urllib.parse import urlparse

REPO_ROOT = Path(__file__).resolve().parents[1]

# (sheet title, table name, row limit) — order defines sheet order.
SHEETS = [
    ("Sources",       "sources",       50),
    ("Experts",       "experts",       20),
    ("Molecules",     "molecules",     20),
    ("Odours",        "odours",        20),
    ("Organizations", "organizations", 20),
]
ORDER_BY = "ORDER BY created_at DESC NULLS LAST, id DESC"
CELL_MAX = 32000   # Excel hard limit is 32767 chars/cell; stay safely under it.


def _database_url() -> str:
    envp = REPO_ROOT / ".env"
    if envp.exists():
        for line in envp.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line.startswith("DATABASE_URL="):
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    import os
    url = os.environ.get("DATABASE_URL")
    if not url:
        raise SystemExit("DATABASE_URL not found (checked meatCODE/.env and environment).")
    return url


def _cell(v):
    """Render a raw DB value into an Excel-safe scalar, unchanged in meaning."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return v
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat(sep=" ")            # tz-aware -> string (openpyxl can't store tz datetimes)
    if isinstance(v, (list, tuple)):
        return ", ".join("" if x is None else str(x) for x in v)
    if isinstance(v, dict):
        s = json.dumps(v, ensure_ascii=False, default=str)
    else:
        s = str(v)
    return s if len(s) <= CELL_MAX else s[:CELL_MAX] + " …[truncated]"


def main() -> int:
    ap = argparse.ArgumentParser(description="Raw Neon → 1 xlsx (5 sheets), newest-first.")
    ap.add_argument("--out", default=None, help="Output .xlsx path (default: data/snapshots/meatcode_snapshot_<UTCstamp>.xlsx)")
    ap.add_argument("--with-ai-review", action="store_true",
                    help="After writing the raw snapshot, append an honest AI Review sheet "
                         "(calls pipeline/ai_review.py in API mode; can take ~60s). Default off "
                         "keeps this export raw-only.")
    args = ap.parse_args()

    import psycopg2
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    url = _database_url()
    taken_utc = _dt.datetime.now(_dt.timezone.utc)
    taken_local = taken_utc.astimezone()
    parsed = urlparse(url)
    origin = f"{parsed.hostname or '?'} / {(parsed.path or '').lstrip('/') or '?'} (Neon Postgres)"

    out_path = Path(args.out) if args.out else (
        REPO_ROOT / "data" / "snapshots" / f"meatcode_snapshot_{taken_utc:%Y%m%d_%H%M}.xlsx"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)

    conn = psycopg2.connect(url)
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; we add our own

    bold = Font(bold=True)
    counts = {}

    for idx, (title, table, limit) in enumerate(SHEETS):
        ws = wb.create_sheet(title=title)
        with conn.cursor() as cur:
            cur.execute(f'SELECT * FROM "{table}" {ORDER_BY} LIMIT %s', (limit,))
            headers = [d[0] for d in cur.description]
            rows = cur.fetchall()
        counts[title] = len(rows)

        start = 1
        if idx == 0:
            # Provenance banner on the FIRST sheet only: exact date+time + origin.
            banner = [
                ("MeatCODE — raw Neon database snapshot", ""),
                ("Snapshot taken (UTC):",   taken_utc.strftime("%Y-%m-%d %H:%M:%S %Z")),
                ("Snapshot taken (local):", taken_local.strftime("%Y-%m-%d %H:%M:%S %Z")),
                ("Origin:", origin),
                ("Contents:", "Sources 50 · Experts 20 · Molecules 20 · Odours 20 · Organizations 20 (newest first, by created_at)"),
            ]
            for r, (a, b) in enumerate(banner, start=1):
                ws.cell(row=r, column=1, value=a).font = bold
                ws.cell(row=r, column=2, value=b)
            start = len(banner) + 2  # blank spacer row, then the table

        # Header row (exact DB column names) + raw data
        for c, h in enumerate(headers, start=1):
            ws.cell(row=start, column=c, value=h).font = bold
        for r, row in enumerate(rows, start=start + 1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=_cell(val))

        ws.freeze_panes = ws.cell(row=start + 1, column=1)
        # light readability: cap column widths
        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 22

    conn.close()
    wb.save(out_path)

    print(f"snapshot written: {out_path}")
    print(f"taken (UTC): {taken_utc:%Y-%m-%d %H:%M:%S}  origin: {origin}")
    print("rows: " + " · ".join(f"{k} {counts[k]}" for k in counts))

    # Optional opinion layer. Kept OUT of the default path so the raw export stays raw,
    # fast and deterministic. The recurring audit adds the review itself (agent-authored,
    # injected) rather than relying on this slower in-process API call.
    if args.with_ai_review:
        try:
            import sys as _sys
            _sys.path.insert(0, str(Path(__file__).resolve().parent))
            import ai_review
            key = ai_review._env_value("ANTHROPIC_API_KEY")
            status = ai_review.review_file(out_path, ai_review.DEFAULT_MODEL, key, force=True)
            print(f"ai review: {status}")
        except Exception as e:  # never let the review break the raw snapshot
            print(f"ai review: SKIPPED ({type(e).__name__}: {e}) — raw snapshot is intact")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
